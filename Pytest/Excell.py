import openpyxl

def no_of_rows(file, sheetname):
    work_book=openpyxl.load_workbook(file)
    sheet=work_book[sheetname]
    return sheet.max_row
def no_of_columns(file, sheetname):
    work_book = openpyxl.load_workbook(file)
    sheet = work_book[sheetname]
    return sheet.max_column
def read_data(file, sheetname, rownumber, columnnumber):
    work_book = openpyxl.load_workbook(file)
    sheet = work_book[sheetname]
    return sheet.cell(rownumber,columnnumber).value
def read_locator(file, sheetname):
    work_book = openpyxl.load_workbook(file)
    sheet = work_book[sheetname]
    rows=no_of_rows(file,sheetname)
    out={}
    for row in range(2,rows+1):
        out[read_data(file,sheetname,row,1)]=(read_data(file,sheetname,row,2),read_data(file,sheetname,row,3))
    return out